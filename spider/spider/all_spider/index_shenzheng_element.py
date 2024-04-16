import sys

sys.path.append("../../")

import requests
import time
import json
from fake_useragent import UserAgent
from spider.utils.ck_utils import client
from datetime import datetime
import os
import openpyxl
import random
from spider.utils.check_table_util import check_row_num

ua = UserAgent()

xls_dir_path = "/work/spider/all_spider/data/"


# 1. request
def requset_index_shenzheng_element(index_code):
    try_count = 0
    while True:
        # 1. url
        url = f"https://www.szse.cn/api/report/ShowReport?SHOWTYPE=xlsx&CATALOGID=1747_zs&TABKEY=tab1&ZSDM={index_code}&random={random.random()}"

        # 2. headers
        proxy_json = requests.get("http://stock_proxy:5010/get").json()
        # print(f"proxy_json: {proxy_json}")
        if proxy_json["https"]:
            # print("没有http代理可用")
            continue

        proxies = {
            "http": f"http://{proxy_json['proxy']}"
        }

        headers = {
            "user-agent": ua.random,
        }
        # print(f"headers: {headers}")
        # print(f"proxies: {proxies}")

        # 3. response
        try_count += 1
        try:
            resp = requests.get(url=url, headers=headers, proxies=proxies, timeout=(1, 10))
        except Exception as e:
            # print(e)
            continue
        if try_count > 10:
            break
        if resp.status_code == 200:
            return resp
        else:
            # print(f"resp {try_count} status code {resp.status_code}")
            pass
    return None


# 2. parse data
def parse_response(resp, index_code):
    if resp is None:
        return False
    else:
        try:
            os.makedirs(xls_dir_path, exist_ok=True)
            with open(f"{xls_dir_path}index_shenzheng_element_{index_code}.xlsx", "wb") as f:
                f.write(resp.content)
            return True
        except Exception as e:
            print(e)
            return False


# 3. insert data
def insert_data_list(index_code, dt):
    values_list = insert_data(index_code)
    for values in values_list:
        sql = f"insert into stock.index_shenzheng_element VALUES ({values}, '{dt}');"
        # print(sql)
        client.client.execute(sql)


def insert_data(index_code):
    workbook = openpyxl.load_workbook(f"{xls_dir_path}index_shenzheng_element_{index_code}.xlsx")
    sheet = workbook[f"指数样本股"]

    values_list = []

    for row_data in sheet.iter_rows(min_row=2, max_row=sheet.max_row, values_only=True):
        count = 0
        values = f" '{index_code}', "
        for row_value in row_data:
            count += 1
            values += f" '{get_str(row_value)}'"
            if count != len(row_data):
                values += ","
        values_list.append(values)
    return values_list


def get_str(v):
    if isinstance(v, int):
        v = str(v)
    elif isinstance(v, float):
        v = str(v)
    elif isinstance(v, str):
        v = v.replace("\'", "")
    else:
        v = json.dumps(v)
    return v


def get_index_list(dt):
    result = client.client.execute(f"select indexCode from stock.index_shenzheng where indexCode != '' and dt='{dt}';")
    result = [t[0] for t in result]
    if result == 0:
        sys.exit(1)
    return result


def run(index_code, dt):
    resp = requset_index_shenzheng_element(index_code)
    if resp is None:
        return
    # print(f"-----------------------------------------------")

    flag = parse_response(resp, index_code)
    # print(f"-----------------------------------------------")

    if flag:
        insert_data_list(index_code, dt)
    else:
        print(f"{index_code} 没有该指数的明细xls")
    # print(f"-----------------------------------------------")


def run_all(dt):
    index_list = get_index_list(dt)
    for index_code in index_list:
        try:
            run(index_code, dt)
            print(f"index_code: {index_code} success")
        except Exception as e:
            print(f"index_code: {index_code} fail {e}")

    check_row_num("index_shenzheng_element", dt)
    print(f"-----------------------------------------------")


if __name__ == '__main__':
    todate = datetime.now()
    dt = todate.strftime('%Y-%m-%d')
    # run("000300", "2023-12-15")
    # run("399001", dt)
    run_all(dt)
