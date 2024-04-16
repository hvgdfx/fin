import sys

sys.path.append("../../")

import requests
import time
import json
from fake_useragent import UserAgent
from spider.utils.ck_utils import client
from datetime import datetime
import os
import openpyxl
import random
from spider.utils.check_table_util import check_row_num


ua = UserAgent()

xls_dir_path = "/work/spider/all_spider/data/"


# 1. request
def requset_index_shenzheng():
    try_count = 0
    while True:
        # 1. url
        url = f"https://www.szse.cn/api/report/ShowReport?SHOWTYPE=xlsx&CATALOGID=1812_zs&TABKEY=tab1&random={random.random()}"

        # 2. headers

        proxy_json = requests.get("http://stock_proxy:5010/get").json()
        print(f"proxy_json: {proxy_json}")
        if proxy_json["https"]:
            print("没有http代理可用")
            continue

        proxies = {
            "http": f"http://{proxy_json['proxy']}"
        }

        headers = {
            "user-agent": ua.random,
            # "Accept-Encoding": "gzip, deflate",
            # "Accept-Language": "en,zh-CN;q=0.9,zh;q=0.8",
            # "Cache-Control": "no-cache",
            # "Connection": "keep-alive",
            # "Host": "query.sse.com.cn",
            # "Pragma": "no-cache",
            # "Referer": "http://www.sse.com.cn/",
        }
        print(f"headers: {headers}")
        print(f"proxies: {proxies}")

        # 3. response

        try_count += 1
        try:
            resp = requests.get(url=url, headers=headers, proxies=proxies, timeout=(1, 10))
        except Exception as e:
            print(e)
            continue
        if try_count > 3:
            break
        if resp.status_code == 200:
            return resp
        else:
            print(f"resp {try_count} status code {resp.status_code}")
    return None


# 2. parse data
def parse_response(resp, dt):
    if resp is None:
        return False
    else:
        try:
            os.makedirs(xls_dir_path, exist_ok=True)
            with open(f"{xls_dir_path}index_shenzheng_{dt}.xlsx", "wb") as f:
                f.write(resp.content)
            return True
        except Exception as e:
            print(e)
            return False


# 3. insert data
def insert_data_list(dt):
    workbook = openpyxl.load_workbook(f"{xls_dir_path}index_shenzheng_{dt}.xlsx")
    sheet = workbook[f"指数列表"]

    values_list = []

    for row_data in sheet.iter_rows(min_row=2, max_row=sheet.max_row, values_only=True):
        count = 0
        values = ""
        for row_value in row_data:
            count += 1
            values += f" '{get_str(row_value)}'"
            if count != len(row_data):
                values += ","
        values_list.append(values)
    print(values_list)

    for values in values_list:
        sql = f"insert into stock.index_shenzheng VALUES ({values}, '{dt}')"
        print(sql)
        client.client.execute(sql)


def get_str(v):
    if isinstance(v, int):
        v = str(v)
    elif isinstance(v, float):
        v = str(v)
    elif isinstance(v, str):
        v = v.replace("\'", "")
    else:
        v = json.dumps(v)
    return v


def run(dt):
    resp = requset_index_shenzheng()
    if resp is None:
        return
    print(f"-----------------------------------------------")

    flag = parse_response(resp, dt)
    print(f"-----------------------------------------------")

    if flag:
        insert_data_list(dt)
    print(f"-----------------------------------------------")

    check_row_num("index_shenzheng", dt)
    print(f"-----------------------------------------------")


if __name__ == '__main__':
    todate = datetime.now()
    dt = todate.strftime('%Y-%m-%d')
    run(dt)
